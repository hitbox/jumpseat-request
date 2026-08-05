"""
Exporting related functionality.
"""
import openpyxl

normal_font = openpyxl.styles.Font(name='Arial')
bold_font = openpyxl.styles.Font(name='Arial', bold=True)

def create_excel(table, jumpseat_requests, metadata=None):
    """
    Return a new openpyxl Workbook object from a list of jumpseat requests.
    """
    wb = openpyxl.Workbook()

    ws = wb.active

    if metadata:
        for key, value in metadata.items():
            # Append key/value
            if isinstance(value, tuple):
                attrname, value = value
                ws.append([key, value])
                cell = ws.cell(row=ws.max_row, column=2)
                setattr(cell, attrname, value)
            else:
                ws.append([key, value])
            # Bold the key cell.
            last_row = ws.max_row
            ws.cell(row=last_row, column=1).font = bold_font
            # Merge value cell so they don't participate in autosizing columns.
            ws.merge_cells(
                start_row = last_row,
                start_column = 2,
                end_row = last_row,
                end_column = len(table.columns),
            )

    # Header
    ws.append([])
    ws.append([column.header for column in table.columns])

    for cell in ws[ws.max_row]:
        cell.font = bold_font

    for request in jumpseat_requests:
        row = [
            None if column.attrname is None else column(request)
            for column in table.columns
        ]
        ws.append(row)

    return wb
